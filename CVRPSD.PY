from __future__ import print_function
from ortools.linear_solver import pywraplp
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import numpy as np
import numpy as numpy  
import math
import scipy.stats as st
import sys

def FDP(Q,media,N):
    ps=np.zeros((N+1))
    for j in range(0,N):  
        p=st.poisson.cdf(Q,media[j], loc=0) 
        ps[j]=p   
        if ps[j]>1.0: 
            p[j]=1.0   
    return ps

def prob(ps,Q,N,dem):
    dmenor=np.zeros((N+1))
    dmayor=np.zeros((N+1))
    pr=np.zeros((N+1))
    for j in range(1,N):
        dmenor[j]=np.sum(dem[0:j+1])
        dmayor[j]=np.sum(dem[0:j+2])  
    pr=FDP(Q,dmenor,N)*(1-FDP(Q,dmayor,N))
    return pr

def costos(N,c):
    cost=np.zeros((N+1))
    sp=np.zeros((N+1))
    for i in range(1,N):
        cost[i]=c[i+1,0]
        for j in range(1,i+1):
            cost[i]=cost[i]+c[j,j+1]
            sp[j]=np.sum(c[0,i+1:N+1])
        cost[i]=cost[j]+sp[j]*2
    return cost

def vesperado(cost,N,dem):
    ar=np.zeros((N+1))
    ar[0]=c[0,1]+(1-ps[0])*(c[1,0]+2*np.sum(c[0]))
    for j in range(1,N):
        ar[j]=cost[j]*pr[j]
    count=0
    for i in range(1,N-1):
        count=count+c[i,i+1]
        c5=count+c[N,0]
    d5=np.sum(dem)
    p5=st.poisson.cdf(Q,d5,loc=0) 
    ar[N]=p5*c5
    return ar

#Computes the distance matrix
def compute_distance_matrix(customers_x, customers_y):
    nb_customers = len(customers_x)
    distance_matrix = [[None for i in range(nb_customers)] for j in range(nb_customers)]
    for i in range(nb_customers):
        distance_matrix[i][i] = 0
        for j in range(nb_customers):
            dist = compute_dist(customers_x[i], customers_x[j], customers_y[i], customers_y[j])
            distance_matrix[i][j] = dist
            distance_matrix[j][i] = dist
    return distance_matrix

# Computes the distances to warehouse
def compute_distance_warehouses(depot_x, depot_y, customers_x, customers_y):
    nb_customers = len(customers_x)
    distance_warehouses = [None] * nb_customers
    for i in range(nb_customers):
        dist = compute_dist(depot_x, customers_x[i], depot_y, customers_y[i])
        distance_warehouses[i] = dist
    return distance_warehouses

def compute_dist(xi, xj, yi, yj):
    exact_dist = math.sqrt(math.pow(xi - xj, 2) + math.pow(yi - yj, 2))
    return int(math.floor(exact_dist + 0.5))

def get_nb_trucks(filename):
    begin = filename.rfind("-k")
    if begin != -1:
        begin += 2
        end = filename.find(".", begin)
        return int(filename[begin:end])
    print ("Error: nb_trucks could not be read from the file name. Enter it from the command line")
    sys.exit(1)

def read_elem(filename):
    with open(filename) as f:
        return [str(elem) for elem in f.read().split()]

def read_input_cvrp(filename):
    file_it = iter(read_elem(filename))
    nb_nodes = 0
    while(1):
        token = next(file_it)
        if token == "DIMENSION":
            next(file_it) # Removes the ":"
            nb_nodes = int(next(file_it))
            nb_customers = nb_nodes - 1
        elif token == "CAPACITY":
            next(file_it) # Removes the ":"
            truck_capacity = int(next(file_it))
        elif token == "EDGE_WEIGHT_TYPE":
            next(file_it) # Removes the ":"
            token = next(file_it)
            if token != "EUC_2D":
                print ("Edge Weight Type " + token + " is not supported (only EUD_2D)")
                sys.exit(1)
        elif token == "NODE_COORD_SECTION":
            break

    customers_x = [None]*(1+nb_customers)
    customers_y = [None]*(1+nb_customers)
    depot_x = 0
    depot_y = 0
    for n in range(nb_nodes):
        node_id = int(next(file_it))
        if node_id != n+1:
            print ("Unexpected index")
            sys.exit(1)
        customers_x[node_id-2] = int(next(file_it))
        customers_y[node_id-2] = int(next(file_it))

    # Compute distance matrix
    distance_matrix = compute_distance_matrix(customers_x, customers_y)
    distance_warehouses = compute_distance_warehouses(depot_x, depot_y, customers_x, customers_y)

    token = next(file_it)
    if token != "DEMAND_SECTION":
        print ("Expected token DEMAND_SECTION")
        sys.exit(1)

    demands = [None]*nb_customers
    for n in range(nb_nodes):
        node_id = int(next(file_it))
        if node_id != n+1:
            print ("Unexpected index")
            sys.exit(1)
        if node_id == 1:
            if  int(next(file_it)) != 0:
                print ("Demand for depot should be 0")
                sys.exit(1)
        else:   
            # -2 because orginal customer indices are in 2..nbNodes
            demands[node_id-2] = int(next(file_it))

    token = next(file_it)
    if token != "DEPOT_SECTION":
        print ("Expected token DEPOT_SECTION")
        sys.exit(1)

    warehouse_id = int(next(file_it))
    if warehouse_id != 1:
        print ("Warehouse id is supposed to be 1")
        sys.exit(1)

    end_of_depot_section = int(next(file_it))
    if end_of_depot_section != -1:
        print ("Expecting only one warehouse, more than one found")
        sys.exit(1)

    return (nb_customers,truck_capacity,distance_matrix,distance_warehouses,demands)

(nclie, Q, c,distance_warehouses,dem)=read_input_cvrp("InstanciasP/P-n101-k4.vrp")
nb_trucks = get_nb_trucks("InstanciasP/P-n101-k4.vrp")
m=nb_trucks
#m=2
#clie=np.array([2,3,4,5,6])
#nodos=np.array([1,2,3,4,5,6])
#N=5
#S=4
#s=np.array([3,4,5,6])
#dem=np.array([2,2,2,2,2])
#Q=60
#c=np.array([[0,2,3,4,3,2],
            #[2,0,4,3,2,2],
            #[3,4,0,3,2,3],
            #[4,3,3,0,3,2],
            #[3,2,2,3,0,2],
            #[2,2,3,2,2,0]])

N=nclie
Q=Q
c=np.array(c)
dem=np.array(dem)
media=dem
clie=np.array(list(range(2,nclie+2)))
nodos=np.array(list(range(1,nclie+2)))
s=np.array(list(range(4,nclie+2)))
S=len(s)
ps=FDP(Q,media,N)
pr=prob(ps,Q,N,dem)
cost=costos(N,c)
ar=vesperado(cost,N,dem)
Qdx=np.sum(ar)

wb=openpyxl.load_workbook('inst.xlsx')
wsdatos=wb['datos']
wsdatos.cell(row=3,column=2).value=Q
wsdatos.cell(row=4,column=2).value=m
wsdatos.cell(row=8,column=2).value=S
for i in range(0,N):
    wsdatos.cell(row=2,column=2+i).value=dem[i]
    wsdatos.cell(row=5,column=2+i).value=clie[i]
for i in range(0,N+1):
    for j in range(0,N+1):
        wsdatos.cell(row=6,column=2+i).value=nodos[i]
        wsdatos.cell(row=9+i,column=2+j).value=c[i,j]
for i in range(0,N-2):
    wsdatos.cell(row=7,column=2+i).value=s[i]
wb.save('inst.xlsx')

modelo=pywraplp.Solver('Vrpsd',pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
x = {}
for i in clie:
    for j in clie:
        if j>i:
            x[(i,j)]=modelo.IntVar(0,1, 'x_i%i_j%i' % (i,j))

for j in clie:
    x[(1,j)]=modelo.IntVar(0,2, 'x_i%i_j%i' % (1,j))

#Restriccion 2:
modelo.Add(modelo.Sum([x[(1,j)] for j in clie])==2*m)  
#Restriccion 3:
for k in clie:
   modelo.Add(modelo.Sum([x[(i,k)] for i in nodos if i<k])+modelo.Sum([x[(k,j)] for j in nodos if k<j])==2)   
#Restriccion 4
modelo.Add(modelo.Sum([x[(i,j)] for i in s for j in s if i<j ])<=S-(np.sum(dem[4:])/(N*Q)))
#Restriccion 5:
for i in nodos:
    for j in nodos:
        if (i>=2) and (i<j):
            modelo.Add(x[(i,j)]<=1)
#Restriccion 6:
for j in nodos:   
    if j>=2:
        modelo.Add(x[(1,j)]<=2)
        
modelo.Minimize(modelo.Sum([x[(i,j)]*c[i-1][j-1] for i in nodos for j in nodos if i<j])+Qdx)

modelo.SetTimeLimit(1200000)
status=modelo.Solve()

if status==pywraplp.Solver.OPTIMAL or status==pywraplp.Solver.FEASIBLE:
    for i in nodos:
        for j in nodos:
            if i<j:
                print('el valor de la x_i%i_j%i es: %d' %(i,j, x[(i,j)].solution_value()))
            
print('Valor objetivo =', modelo.Objective().Value())
